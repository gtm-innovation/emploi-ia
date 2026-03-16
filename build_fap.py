"""
Aggregate ROME-level AI exposure scores to FAP228 occupation groups.

Uses:
- ROME→FAP341 mapping (Dares_Table_passage_ROME_Qualif_to_FAP2021)
- FAP341→FAP228 hierarchy (Dares_Arborescence_FAP2021)
- ROME scores (scores.json)
- BMO 2025 hiring projections (bmo_2025.xlsx, uses FAP228 codes)

Outputs site/data.json with FAP228-level entries.

Usage:
    uv run python build_fap.py
"""

import csv
import json
import openpyxl
from collections import defaultdict


def slugify(text):
    import re
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")


def main():
    # 1. Load FAP341→FAP228 hierarchy
    wb = openpyxl.load_workbook(
        "Dares_Arborescence_FAP2021.xlsx", read_only=True
    )
    ws = wb["niveaux_emboités"]
    fap341_to_228 = {}
    fap228_names = {}
    fap228_to_86 = {}
    fap86_names = {}
    fap86_to_22 = {}
    fap22_names = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code341, name341, code228, name228, code86, name86, code22, name22 = row
        fap341_to_228[code341] = code228
        fap228_names[code228] = name228
        fap228_to_86[code228] = code86
        fap86_names[code86] = name86
        fap86_to_22[code86] = code22
        fap22_names[code22] = name22

    print(f"FAP hierarchy: {len(fap228_names)} FAP228 groups")

    # 2. Load ROME→FAP341 mapping
    rome_to_fap228 = defaultdict(set)
    with open(
        "Dares_Table_passage_ROME_Qualif_to_FAP2021_pour_programme.csv",
        encoding="latin-1",
    ) as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rome = row["ROME"]
            fap341 = row["FAP341"]
            fap228 = fap341_to_228.get(fap341, fap341)
            rome_to_fap228[rome].add(fap228)

    print(f"ROME→FAP228 mapping: {len(rome_to_fap228)} ROME codes")

    # 3. Load ROME scores
    with open("scores.json") as f:
        scores_list = json.load(f)
    scores = {s["slug"]: s for s in scores_list}

    # 4. Load occupations (for ROME code→slug mapping)
    with open("occupations.json") as f:
        occupations = json.load(f)
    rome_to_occ = {o["code_rome"]: o for o in occupations}

    # 5. Aggregate scores to FAP228
    fap228_scores = defaultdict(list)
    fap228_rome_list = defaultdict(list)

    for occ in occupations:
        code = occ["code_rome"]
        slug = occ["slug"]
        score_entry = scores.get(slug)
        if not score_entry or "exposure" not in score_entry:
            continue

        fap_codes = rome_to_fap228.get(code, set())
        for fap228 in fap_codes:
            fap228_scores[fap228].append(score_entry["exposure"])
            fap228_rome_list[fap228].append({
                "title": occ["title"],
                "code_rome": code,
                "exposure": score_entry["exposure"],
            })

    # For ROME codes not in mapping, try prefix matching
    unmapped = 0
    for occ in occupations:
        code = occ["code_rome"]
        if code not in rome_to_fap228:
            unmapped += 1

    print(f"Mapped ROME→FAP228: {len(rome_to_fap228)}")
    print(f"Unmapped ROME codes: {unmapped}")

    # 6. Load BMO 2025 hiring projections (aggregated nationally)
    wb_bmo = openpyxl.load_workbook("bmo_2025.xlsx", read_only=True)
    ws_bmo = wb_bmo["BMO_2025_open_data"]
    bmo = defaultdict(int)
    for row in ws_bmo.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        code = row[1]
        val = row[11]  # met column
        if val and str(val) != "*":
            try:
                bmo[code] += int(val)
            except (ValueError, TypeError):
                pass

    print(f"BMO 2025: {len(bmo)} FAP codes, {sum(bmo.values()):,} projets")

    # 7. Load DARES median salary data (2017-2019)
    salary_by_code = {}
    with open("dares_salaire_median.csv", encoding="latin-1") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            code = row["Code Fap"]
            year = row.get("Ann\xe9e", "")
            sal = row.get("Salaire m\xe9dian (en \x80)", "")
            if "2017-2019" in year and sal and sal != "nd":
                salary_by_code[code] = int(sal)

    def get_salary(code228):
        """Map FAP2021 code to old FAP salary data (X→Z substitution)."""
        old = code228.replace("X", "Z")
        # Try FAP228 level
        if old in salary_by_code:
            return salary_by_code[old]
        # Try FAP86 level (3 chars)
        old86 = old[:3]
        if old86 in salary_by_code:
            return salary_by_code[old86]
        # Try FAP22 level (1 char)
        if old[0] in salary_by_code:
            return salary_by_code[old[0]]
        return None

    print(f"DARES salary: {len(salary_by_code)} codes loaded")

    # 8. Build site data
    data = []
    scored_count = 0

    for code228, name228 in sorted(fap228_names.items()):
        score_list = fap228_scores.get(code228, [])
        rome_details = fap228_rome_list.get(code228, [])
        hiring = bmo.get(code228, 0)

        if not score_list:
            avg_exposure = None
        else:
            avg_exposure = round(sum(score_list) / len(score_list), 1)
            scored_count += 1

        code86 = fap228_to_86.get(code228, "")
        category = fap86_names.get(code86, "Autre")

        # Build rationale from top-scoring ROME occupations
        rome_details_sorted = sorted(
            rome_details, key=lambda x: -x["exposure"]
        )
        top_exposed = [
            r["title"] for r in rome_details_sorted[:3]
            if r["exposure"] >= 7
        ]
        low_exposed = [
            r["title"] for r in rome_details_sorted[-3:]
            if r["exposure"] <= 3
        ]

        rationale_parts = []
        if top_exposed:
            rationale_parts.append(
                f"Métiers les plus exposés : {', '.join(top_exposed)}."
            )
        if low_exposed:
            rationale_parts.append(
                f"Métiers les moins exposés : {', '.join(low_exposed)}."
            )
        rationale_parts.append(
            f"Score moyen sur {len(score_list)} métiers ROME."
        )

        entry = {
            "title": name228,
            "slug": slugify(name228),
            "category": category,
            "code_fap": code228,
            "jobs": hiring,
            "salary": get_salary(code228),
            "exposure": avg_exposure if avg_exposure is not None else 0,
            "exposure_rationale": " ".join(rationale_parts),
            "rome_count": len(score_list),
        }
        data.append(entry)

    # Filter out entries with no hiring and no scores
    data = [d for d in data if d["jobs"] > 0 or d["rome_count"] > 0]

    with open("site/data.json", "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    total_hiring = sum(d["jobs"] for d in data)
    print(f"\nWrote {len(data)} FAP228 groups to site/data.json")
    print(f"  {scored_count} scored, {len(data) - scored_count} unscored")
    print(f"  Total projets de recrutement : {total_hiring:,}")

    # Distribution
    exposures = [d["exposure"] for d in data if d["exposure"] > 0]
    if exposures:
        avg = sum(exposures) / len(exposures)
        print(f"  Exposition moyenne : {avg:.1f}")


if __name__ == "__main__":
    main()
